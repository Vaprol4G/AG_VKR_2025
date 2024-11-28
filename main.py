import openpyxl
from docxtpl import DocxTemplate
import xml.etree.ElementTree as ET


# Функция для генерации документа по шаблону
def DocXCreate(context):
    context_correct = {}
    count = 0
    for key in context:
        context_correct[('Data_' + str(count))] = context[key]
        count += 1
    doc = DocxTemplate("sample.docx")
    doc.render(context_correct)
    # print(context_correct)
    # Название документа
    doc.save(context['Фамилия'] + ' ' + context['Имя'] + ' ' + context['Отчество'] + '.docx')


XL_File_Name = 'XL.xlsx'
XML_File_Name = 'output.xml'

# Чтение данных из таблицы
# Формирование шаблона для словаря
XL = openpyxl.load_workbook(XL_File_Name)
XL_SD = {XL.active['A1'].value: '', XL.active['B1'].value: '', XL.active['C1'].value: '',
         XL.active['D1'].value: '', XL.active['E1'].value: '', XL.active['F1'].value: '',
         XL.active['G1'].value: '', XL.active['H1'].value: '', XL.active['I1'].value: '',
         XL.active['J1'].value: '', XL.active['K1'].value: '', XL.active['L1'].value: '',
         XL.active['M1'].value: '', XL.active['N1'].value: '', XL.active['O1'].value: '', XL.active['P1'].value: ''}

# Заполнение списка со словорями данными из таблицы
XL_Data = []
XL_Count = 0
while True:
    if XL.active.cell(row=(XL_Count + 2), column=1).value is None:
        break
    else:
        XL_Data.append(XL_SD.copy())
        for i in range(16):
            XL_Data[XL_Count][XL.active.cell(row=1, column=(i + 1)).value] = XL.active.cell(row=(XL_Count + 2),
                                                                                            column=(i + 1)).value
        # print(XL_Data[XL_Count])
        XL_Count += 1

# Вызов функции (для теста только один)
DocXCreate(XL_Data[0])

# Создание XML
root = ET.Element("RegistryRecords")

for data in XL_Data:
    registry_record = ET.SubElement(root, "RegistryRecord")

    worker = ET.SubElement(registry_record, "Worker")
    ET.SubElement(worker, "LastName").text = data.get("Фамилия", "")
    ET.SubElement(worker, "FirstName").text = data.get("Имя", "")
    ET.SubElement(worker, "MiddleName").text = data.get("Отчество", "")
    ET.SubElement(worker, "Snils").text = data.get("СНИЛС", "")
    ET.SubElement(worker, "Position").text = data.get("Должность", "")
    ET.SubElement(worker, "EmployerInn").text = data.get("ИНН работодателя", "")
    ET.SubElement(worker, "EmployerTitle").text = data.get("Наименование работодателя", "")

    organization = ET.SubElement(registry_record, "Organization")
    ET.SubElement(organization, "Inn").text = data.get("ИНН организации проводившей обучение", "")
    ET.SubElement(organization, "Title").text = data.get("Наименование организации проводившей обучение", "")

    result = data.get("Результат", "").strip()
    is_passed = "true" if result == "Успешно" else "false"

    test = ET.SubElement(registry_record, "Test", isPassed=is_passed)
    ET.SubElement(test, "Date").text = data.get("Дата проверки знаний", "")
    ET.SubElement(test, "ProtocolNumber").text = data.get("Номер протокола", "")
    ET.SubElement(test, "LearnProgramTitle").text = data.get("Программа обучения", "")

    log = ET.SubElement(registry_record, "Log")
    ET.SubElement(log, "LogNumber").text = data.get("Номер в реестре", "")
    ET.SubElement(log, "PackageId").text = data.get("Id пакета", "")
    ET.SubElement(log, "LogEntryDate").text = data.get("Дата внесения в реестр", "")

tree = ET.ElementTree(root)
tree.write(XML_File_Name, encoding="utf-8", xml_declaration=True)
